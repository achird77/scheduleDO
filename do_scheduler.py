#!/usr/bin/env python3
"""
Duty Officer Schedule Planner
=============================

A single-file, fully-customizable Duty Officer (DO) rostering tool with a GUI.

Core rules implemented (all configurable):
  * 12-hour watch shifts: one DAY watch + one NIGHT watch required every day.
  * 4-6 (configurable) duty officers; monthly hours kept ~equal.
  * Optional 1-2 officers on an 8h Mon-Fri "day work" shift, rotated weekly.
  * Weekends: the Saturday DAY person also works Sunday DAY; same for NIGHT.
  * Weekend duties spread as evenly as possible.
  * Minimum 2 consecutive watch shifts; maximum 3 consecutive.
  * No DAY->NIGHT or NIGHT->DAY on consecutive days (rest in between).
  * Day & night shift counts kept ~equal across officers.
  * Re-generate button (new random valid roster each time).
  * Export to Excel (.xlsx) or CSV.

Solver: Google OR-Tools CP-SAT (hard safety constraints + soft fairness).
GUI:    Tkinter (standard library).

Run:
    python3 do_scheduler.py            # launch GUI
    python3 do_scheduler.py --selftest # headless engine validation
"""

import sys
import csv
import random
import datetime as dt
from dataclasses import dataclass, field
from typing import Optional

# ----------------------------------------------------------------------------
# Optional dependencies (checked lazily so the GUI can show friendly messages)
# ----------------------------------------------------------------------------
try:
    from ortools.sat.python import cp_model
    HAVE_ORTOOLS = True
except Exception:
    HAVE_ORTOOLS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False


WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Assignment codes used throughout
OFF = "OFF"
DAY = "DAY"
NIGHT = "NIGHT"
DAYWORK = "DAYWORK"


# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------
@dataclass
class ShiftDef:
    label: str
    time_text: str
    hours: int


@dataclass
class ScheduleConfig:
    officers: list = field(default_factory=lambda: ["DO 1", "DO 2", "DO 3", "DO 4"])
    weeks: int = 4
    start_date: dt.date = field(default_factory=lambda: ScheduleConfig._next_monday())

    # Shift definitions
    day_shift: ShiftDef = field(default_factory=lambda: ShiftDef("DAY", "1000-2200", 12))
    night_shift: ShiftDef = field(default_factory=lambda: ShiftDef("NIGHT", "2200-1000", 12))
    daywork_shift: ShiftDef = field(default_factory=lambda: ShiftDef("DAY WORK", "0800-1700", 8))

    # Number of 8h Mon-Fri day-work officers per week (0, 1 or 2), rotated weekly
    daywork_per_week: int = 0

    # Block rules (apply to 12h watch shifts)
    min_consecutive: int = 2
    max_consecutive: int = 3
    rest_days_after_block: int = 1  # minimum OFF days required after a watch block

    # Off days required when changing shift TYPE
    switch_daynight_off_days: int = 2    # day<->night change needs >= this many OFF days
    switch_workhours_off_days: int = 1   # 12h<->8h change needs >= this many OFF days

    # Whether the min-consecutive rule is hard (True) or strongly preferred (False)
    enforce_min_consecutive: bool = True

    # Fairness weights (higher = more important to equalise)
    w_hours: int = 100
    w_daynight_self: int = 40   # balance of own day vs night counts
    w_daynight_spread: int = 20  # spread of day/night counts across officers
    w_weekend: int = 30
    w_daywork_rotation: int = 25

    # Solver
    time_limit_s: int = 20

    @staticmethod
    def _next_monday(from_date: Optional[dt.date] = None) -> dt.date:
        d = from_date or dt.date.today()
        return d + dt.timedelta(days=(7 - d.weekday()) % 7 if d.weekday() != 0 else 0)

    @property
    def num_days(self) -> int:
        return self.weeks * 7

    def date_for(self, t: int) -> dt.date:
        return self.start_date + dt.timedelta(days=t)

    def is_weekday(self, t: int) -> bool:
        return self.date_for(t).weekday() < 5  # Mon-Fri

    def is_saturday(self, t: int) -> bool:
        return self.date_for(t).weekday() == 5

    def is_sunday(self, t: int) -> bool:
        return self.date_for(t).weekday() == 6


# ----------------------------------------------------------------------------
# Result container
# ----------------------------------------------------------------------------
@dataclass
class ScheduleResult:
    status: str                         # "OPTIMAL" / "FEASIBLE" / "INFEASIBLE" / "ERROR"
    message: str
    # grid[officer_index][day_index] -> assignment code (OFF/DAY/NIGHT/DAYWORK)
    grid: list = field(default_factory=list)
    config: Optional[ScheduleConfig] = None

    @property
    def ok(self) -> bool:
        return self.status in ("OPTIMAL", "FEASIBLE")

    def hours_for(self, o: int) -> int:
        cfg = self.config
        h = 0
        for code in self.grid[o]:
            if code == DAY:
                h += cfg.day_shift.hours
            elif code == NIGHT:
                h += cfg.night_shift.hours
            elif code == DAYWORK:
                h += cfg.daywork_shift.hours
        return h

    def count_for(self, o: int, code: str) -> int:
        return sum(1 for c in self.grid[o] if c == code)

    def weekend_count_for(self, o: int) -> int:
        cfg = self.config
        n = 0
        for t in range(cfg.num_days):
            if (cfg.is_saturday(t) or cfg.is_sunday(t)) and self.grid[o][t] in (DAY, NIGHT):
                n += 1
        return n


# ----------------------------------------------------------------------------
# Scheduling engine (CP-SAT)
# ----------------------------------------------------------------------------
def generate_schedule(cfg: ScheduleConfig, seed: Optional[int] = None) -> ScheduleResult:
    if not HAVE_ORTOOLS:
        return ScheduleResult("ERROR",
                              "Google OR-Tools is not installed.\n"
                              "Install it with:  pip install ortools")
    if seed is None:
        seed = random.randint(1, 2_000_000_000)
    rng = random.Random(seed)

    N = len(cfg.officers)
    T = cfg.num_days
    W = cfg.weeks
    if N < 2:
        return ScheduleResult("ERROR", "Need at least 2 duty officers.")

    m = cp_model.CpModel()

    # Decision variables --------------------------------------------------
    d = {(o, t): m.NewBoolVar(f"day_{o}_{t}") for o in range(N) for t in range(T)}
    n = {(o, t): m.NewBoolVar(f"night_{o}_{t}") for o in range(N) for t in range(T)}
    w = {(o, t): m.NewBoolVar(f"work_{o}_{t}") for o in range(N) for t in range(T)}
    dwk = {(o, wk): m.NewBoolVar(f"dwweek_{o}_{wk}") for o in range(N) for wk in range(W)}

    def week_of(t):
        return t // 7

    # At most one assignment per officer per day
    for o in range(N):
        for t in range(T):
            m.AddAtMostOne([d[o, t], n[o, t], w[o, t]])

    # Day-work weekly linkage --------------------------------------------
    for o in range(N):
        for wk in range(W):
            for t in range(wk * 7, min((wk + 1) * 7, T)):
                if cfg.is_weekday(t):
                    m.Add(w[o, t] == dwk[o, wk])
                else:
                    m.Add(w[o, t] == 0)
                # No watches during a day-work week
                m.Add(d[o, t] + n[o, t] <= 1 - dwk[o, wk])

    # Coverage ------------------------------------------------------------
    for t in range(T):
        m.Add(sum(d[o, t] for o in range(N)) == 1)
        m.Add(sum(n[o, t] for o in range(N)) == 1)
    for wk in range(W):
        m.Add(sum(dwk[o, wk] for o in range(N)) == cfg.daywork_per_week)

    # Weekend: Sat person == Sun person for each watch -------------------
    for t in range(T - 1):
        if cfg.is_saturday(t) and cfg.is_sunday(t + 1):
            for o in range(N):
                m.Add(d[o, t] == d[o, t + 1])
                m.Add(n[o, t] == n[o, t + 1])

    # Shift-change rest rules ------------------------------------------
    # (a) Changing day<->night requires >= switch_daynight_off_days OFF days:
    #     forbid a NIGHT within that many days after any DAY (and vice versa).
    g_dn = max(1, cfg.switch_daynight_off_days)
    for o in range(N):
        for t in range(T):
            for g in range(1, g_dn + 1):
                if t + g < T:
                    m.Add(d[o, t] + n[o, t + g] <= 1)
                    m.Add(n[o, t] + d[o, t + g] <= 1)

    # (b) Changing 12h watch <-> 8h day-work requires >= switch_workhours_off_days OFF days.
    g_wh = max(1, cfg.switch_workhours_off_days)
    for o in range(N):
        for t in range(T):
            for g in range(1, g_wh + 1):
                if t + g < T:
                    m.Add((d[o, t] + n[o, t]) + w[o, t + g] <= 1)
                    m.Add(w[o, t] + (d[o, t + g] + n[o, t + g]) <= 1)

    # Max consecutive watch days -----------------------------------------
    mc = cfg.max_consecutive
    for o in range(N):
        for t in range(T - mc):
            m.Add(sum(d[o, t + k] + n[o, t + k] for k in range(mc + 1)) <= mc)

    # watch indicator var
    watch = {(o, t): m.NewBoolVar(f"watch_{o}_{t}") for o in range(N) for t in range(T)}
    for o in range(N):
        for t in range(T):
            m.Add(watch[o, t] == d[o, t] + n[o, t])

    # Minimum consecutive (no isolated single watch day) -----------------
    iso_pen = []
    for o in range(N):
        for t in range(T):
            neighbors = []
            if t - 1 >= 0:
                neighbors.append(watch[o, t - 1])
            if t + 1 < T:
                neighbors.append(watch[o, t + 1])
            if cfg.enforce_min_consecutive and cfg.min_consecutive >= 2:
                # watch[t] must have at least one watching neighbor
                m.Add(watch[o, t] <= sum(neighbors))
            else:
                isov = m.NewBoolVar(f"iso_{o}_{t}")
                # isov >= watch[t] - sum(neighbors)
                m.Add(isov >= watch[o, t] - sum(neighbors))
                iso_pen.append(isov)

    # Rest days after a block (only needed if > 1, since >=1 is automatic)
    if cfg.rest_days_after_block >= 2:
        R = cfg.rest_days_after_block
        for o in range(N):
            for t in range(T - 1):
                # block end at t: watch[t]=1, watch[t+1]=0  -> next R days off
                for r in range(2, R + 1):
                    if t + r < T:
                        # if watch[t]=1 and watch[t+1]=0 then watch[t+r]=0
                        m.Add(watch[o, t + r] <= watch[o, t + 1] + (1 - watch[o, t]))

    # ------------------------------------------------------------------
    # Fairness objective terms
    # ------------------------------------------------------------------
    obj_terms = []

    # Total hours per officer
    hours = {}
    max_h = T * 12
    for o in range(N):
        hv = m.NewIntVar(0, max_h, f"hours_{o}")
        m.Add(hv == sum(d[o, t] * cfg.day_shift.hours
                        + n[o, t] * cfg.night_shift.hours
                        + w[o, t] * cfg.daywork_shift.hours for t in range(T)))
        hours[o] = hv
    hmax = m.NewIntVar(0, max_h, "hmax")
    hmin = m.NewIntVar(0, max_h, "hmin")
    m.AddMaxEquality(hmax, [hours[o] for o in range(N)])
    m.AddMinEquality(hmin, [hours[o] for o in range(N)])
    hspread = m.NewIntVar(0, max_h, "hspread")
    m.Add(hspread == hmax - hmin)
    obj_terms.append(cfg.w_hours * hspread)

    # Day & night counts per officer
    day_tot, night_tot = {}, {}
    for o in range(N):
        dv = m.NewIntVar(0, T, f"daytot_{o}")
        nv = m.NewIntVar(0, T, f"nighttot_{o}")
        m.Add(dv == sum(d[o, t] for t in range(T)))
        m.Add(nv == sum(n[o, t] for t in range(T)))
        day_tot[o], night_tot[o] = dv, nv

    # Self day/night balance: |day - night| per officer
    for o in range(N):
        diff = m.NewIntVar(-T, T, f"dndiff_{o}")
        m.Add(diff == day_tot[o] - night_tot[o])
        absdiff = m.NewIntVar(0, T, f"dnabs_{o}")
        m.AddAbsEquality(absdiff, diff)
        obj_terms.append(cfg.w_daynight_self * absdiff)

    # Spread of day counts and night counts across officers
    for grp, name in [(day_tot, "day"), (night_tot, "night")]:
        gmax = m.NewIntVar(0, T, f"{name}max")
        gmin = m.NewIntVar(0, T, f"{name}min")
        m.AddMaxEquality(gmax, [grp[o] for o in range(N)])
        m.AddMinEquality(gmin, [grp[o] for o in range(N)])
        sp = m.NewIntVar(0, T, f"{name}spread")
        m.Add(sp == gmax - gmin)
        obj_terms.append(cfg.w_daynight_spread * sp)

    # Weekend fairness
    wk_tot = {}
    weekend_days = [t for t in range(T) if cfg.is_saturday(t) or cfg.is_sunday(t)]
    for o in range(N):
        wv = m.NewIntVar(0, len(weekend_days), f"wkend_{o}")
        m.Add(wv == sum(watch[o, t] for t in weekend_days))
        wk_tot[o] = wv
    if weekend_days:
        wkmax = m.NewIntVar(0, len(weekend_days), "wkmax")
        wkmin = m.NewIntVar(0, len(weekend_days), "wkmin")
        m.AddMaxEquality(wkmax, [wk_tot[o] for o in range(N)])
        m.AddMinEquality(wkmin, [wk_tot[o] for o in range(N)])
        wksp = m.NewIntVar(0, len(weekend_days), "wkspread")
        m.Add(wksp == wkmax - wkmin)
        obj_terms.append(cfg.w_weekend * wksp)

    # Day-work rotation fairness (even number of dw-weeks per officer)
    if cfg.daywork_per_week > 0:
        dw_tot = {}
        for o in range(N):
            dv = m.NewIntVar(0, W, f"dwtot_{o}")
            m.Add(dv == sum(dwk[o, wk] for wk in range(W)))
            dw_tot[o] = dv
        dwmax = m.NewIntVar(0, W, "dwmax")
        dwmin = m.NewIntVar(0, W, "dwmin")
        m.AddMaxEquality(dwmax, [dw_tot[o] for o in range(N)])
        m.AddMinEquality(dwmin, [dw_tot[o] for o in range(N)])
        dwsp = m.NewIntVar(0, W, "dwspread")
        m.Add(dwsp == dwmax - dwmin)
        obj_terms.append(cfg.w_daywork_rotation * dwsp)
        # Discourage same officer doing day-work two consecutive weeks
        for o in range(N):
            for wk in range(W - 1):
                same = m.NewBoolVar(f"dwsame_{o}_{wk}")
                m.Add(same >= dwk[o, wk] + dwk[o, wk + 1] - 1)
                obj_terms.append(cfg.w_daywork_rotation * same)

    # Soft isolated-day penalty (only when min-consecutive is not hard)
    for v in iso_pen:
        obj_terms.append(1000 * v)

    # Random tie-breaking so "Re-generate" yields different valid rosters
    for o in range(N):
        for t in range(T):
            obj_terms.append(rng.randint(0, 3) * d[o, t])
            obj_terms.append(rng.randint(0, 3) * n[o, t])

    m.Minimize(sum(obj_terms))

    # Solve ---------------------------------------------------------------
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(cfg.time_limit_s)
    solver.parameters.random_seed = seed % 2_000_000_000
    solver.parameters.num_search_workers = 8
    status = solver.Solve(m)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return ScheduleResult(
            "INFEASIBLE",
            "No valid roster exists for these settings.\n\n"
            "Try one of the following:\n"
            "  • Add another duty officer\n"
            "  • Reduce day-work officers per week\n"
            "  • Relax the consecutive-shift rules\n"
            "  • Disable 'enforce minimum consecutive shifts'",
            config=cfg)

    grid = [[OFF] * T for _ in range(N)]
    for o in range(N):
        for t in range(T):
            if solver.Value(d[o, t]):
                grid[o][t] = DAY
            elif solver.Value(n[o, t]):
                grid[o][t] = NIGHT
            elif solver.Value(w[o, t]):
                grid[o][t] = DAYWORK

    status_name = "OPTIMAL" if status == cp_model.OPTIMAL else "FEASIBLE"
    return ScheduleResult(status_name,
                          f"Roster generated ({status_name.lower()}).",
                          grid=grid, config=cfg)


# ----------------------------------------------------------------------------
# Validation (used by self-test and as a safety net)
# ----------------------------------------------------------------------------
def validate(res: ScheduleResult) -> list:
    """Return a list of rule-violation strings (empty = perfectly valid)."""
    cfg = res.config
    N, T = len(cfg.officers), cfg.num_days
    problems = []

    # Coverage
    for t in range(T):
        dc = sum(1 for o in range(N) if res.grid[o][t] == DAY)
        nc = sum(1 for o in range(N) if res.grid[o][t] == NIGHT)
        if dc != 1:
            problems.append(f"Day {t}: {dc} DAY watchers (need 1)")
        if nc != 1:
            problems.append(f"Day {t}: {nc} NIGHT watchers (need 1)")

    # Transitions + blocks
    g_dn = max(1, cfg.switch_daynight_off_days)
    g_wh = max(1, cfg.switch_workhours_off_days)
    for o in range(N):
        row = res.grid[o]
        for t in range(T):
            # day<->night spacing
            for g in range(1, g_dn + 1):
                if t + g < T:
                    if row[t] == DAY and row[t + g] == NIGHT:
                        problems.append(f"{cfg.officers[o]}: DAY->NIGHT only {g-1} day(s) off (day {t})")
                    if row[t] == NIGHT and row[t + g] == DAY:
                        problems.append(f"{cfg.officers[o]}: NIGHT->DAY only {g-1} day(s) off (day {t})")
            # 12h<->8h spacing
            for g in range(1, g_wh + 1):
                if t + g < T:
                    a, b = row[t], row[t + g]
                    if (a in (DAY, NIGHT) and b == DAYWORK) or (a == DAYWORK and b in (DAY, NIGHT)):
                        problems.append(f"{cfg.officers[o]}: 12h<->8h only {g-1} day(s) off (day {t})")
        # block lengths for watches
        run = 0
        for t in range(T + 1):
            iswatch = t < T and row[t] in (DAY, NIGHT)
            if iswatch:
                run += 1
            else:
                if run > 0:
                    if cfg.enforce_min_consecutive and run < cfg.min_consecutive:
                        problems.append(f"{cfg.officers[o]}: block of {run} (<{cfg.min_consecutive}) ending day {t-1}")
                    if run > cfg.max_consecutive:
                        problems.append(f"{cfg.officers[o]}: block of {run} (>{cfg.max_consecutive}) ending day {t-1}")
                run = 0

    # Weekend same-person
    for t in range(T - 1):
        if cfg.is_saturday(t) and cfg.is_sunday(t + 1):
            for o in range(N):
                if res.grid[o][t] == DAY and res.grid[o][t + 1] != DAY:
                    problems.append(f"{cfg.officers[o]}: Sat DAY not matched on Sun (day {t})")
                if res.grid[o][t] == NIGHT and res.grid[o][t + 1] != NIGHT:
                    problems.append(f"{cfg.officers[o]}: Sat NIGHT not matched on Sun (day {t})")
    return problems


# ----------------------------------------------------------------------------
# Human-readable rule list
# ----------------------------------------------------------------------------
def rules_text(cfg: ScheduleConfig) -> list:
    ds_, ns_, ws_ = cfg.day_shift, cfg.night_shift, cfg.daywork_shift
    lines = [
        f"1. Daily coverage: exactly one {ds_.label} watch ({ds_.time_text}, {ds_.hours}h) "
        f"and one {ns_.label} watch ({ns_.time_text}, {ns_.hours}h) every day.",
        f"2. {len(cfg.officers)} duty officers; total monthly hours kept as equal as possible.",
    ]
    if cfg.daywork_per_week > 0:
        lines.append(
            f"3. {cfg.daywork_per_week} officer(s) per week on the 8h day-work shift "
            f"({ws_.label} {ws_.time_text}, {ws_.hours}h), Mon-Fri, rotated weekly.")
    else:
        lines.append("3. No 8h day-work shift in use.")
    lines += [
        "4. Weekends: the officer on Saturday also works Sunday "
        "(one officer for the day watch, one for the night watch).",
        "5. Weekend duties are spread across officers as evenly as possible.",
        f"6. Watch blocks: at least {cfg.min_consecutive} and at most "
        f"{cfg.max_consecutive} consecutive watch days"
        + ("." if cfg.enforce_min_consecutive else " (minimum is preferred, not strict)."),
        f"7. Switching between day and night watches requires at least "
        f"{cfg.switch_daynight_off_days} day(s) off in between.",
        f"8. Switching between a 12h watch and the 8h day-work shift requires at least "
        f"{cfg.switch_workhours_off_days} day(s) off in between.",
        f"9. At least {cfg.rest_days_after_block} rest day(s) after each watch block.",
        "10. Day and night shift counts kept as equal as possible "
        "for each officer and across the team.",
    ]
    return lines


# ----------------------------------------------------------------------------
# Save / load a session (settings + current roster) as JSON
# ----------------------------------------------------------------------------
def config_to_dict(cfg: ScheduleConfig) -> dict:
    return {
        "officers": cfg.officers,
        "weeks": cfg.weeks,
        "start_date": cfg.start_date.isoformat(),
        "day_shift": [cfg.day_shift.label, cfg.day_shift.time_text, cfg.day_shift.hours],
        "night_shift": [cfg.night_shift.label, cfg.night_shift.time_text, cfg.night_shift.hours],
        "daywork_shift": [cfg.daywork_shift.label, cfg.daywork_shift.time_text, cfg.daywork_shift.hours],
        "daywork_per_week": cfg.daywork_per_week,
        "min_consecutive": cfg.min_consecutive,
        "max_consecutive": cfg.max_consecutive,
        "rest_days_after_block": cfg.rest_days_after_block,
        "switch_daynight_off_days": cfg.switch_daynight_off_days,
        "switch_workhours_off_days": cfg.switch_workhours_off_days,
        "enforce_min_consecutive": cfg.enforce_min_consecutive,
        "w_hours": cfg.w_hours,
        "w_daynight_self": cfg.w_daynight_self,
        "w_daynight_spread": cfg.w_daynight_spread,
        "w_weekend": cfg.w_weekend,
        "w_daywork_rotation": cfg.w_daywork_rotation,
        "time_limit_s": cfg.time_limit_s,
    }


def config_from_dict(d: dict) -> ScheduleConfig:
    return ScheduleConfig(
        officers=list(d["officers"]),
        weeks=int(d["weeks"]),
        start_date=dt.date.fromisoformat(d["start_date"]),
        day_shift=ShiftDef(*d["day_shift"]),
        night_shift=ShiftDef(*d["night_shift"]),
        daywork_shift=ShiftDef(*d["daywork_shift"]),
        daywork_per_week=int(d["daywork_per_week"]),
        min_consecutive=int(d["min_consecutive"]),
        max_consecutive=int(d["max_consecutive"]),
        rest_days_after_block=int(d["rest_days_after_block"]),
        switch_daynight_off_days=int(d.get("switch_daynight_off_days", 2)),
        switch_workhours_off_days=int(d.get("switch_workhours_off_days", 1)),
        enforce_min_consecutive=bool(d["enforce_min_consecutive"]),
        w_hours=int(d["w_hours"]),
        w_daynight_self=int(d["w_daynight_self"]),
        w_daynight_spread=int(d["w_daynight_spread"]),
        w_weekend=int(d["w_weekend"]),
        w_daywork_rotation=int(d["w_daywork_rotation"]),
        time_limit_s=int(d["time_limit_s"]),
    )


def save_session(path: str, cfg: ScheduleConfig, res: Optional[ScheduleResult]):
    import json
    data = {"config": config_to_dict(cfg),
            "grid": res.grid if (res and res.ok) else None,
            "status": res.status if res else None}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def load_session(path: str):
    import json
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    cfg = config_from_dict(data["config"])
    res = None
    if data.get("grid"):
        res = ScheduleResult("FEASIBLE", "Loaded from file.", grid=data["grid"], config=cfg)
    return cfg, res


# ----------------------------------------------------------------------------
# Export
# ----------------------------------------------------------------------------
def _cell_text(res: ScheduleResult, o: int, t: int) -> str:
    code = res.grid[o][t]
    cfg = res.config
    if code == DAY:
        return cfg.day_shift.time_text
    if code == NIGHT:
        return cfg.night_shift.time_text
    if code == DAYWORK:
        return cfg.daywork_shift.time_text
    return "OFF"


def export_csv(res: ScheduleResult, path: str):
    cfg = res.config
    T = cfg.num_days
    with open(path, "w", newline="", encoding="utf-8") as f:
        wr = csv.writer(f)
        wr.writerow(["DUTY OFFICER SCHEDULE", f"Start: {cfg.start_date.isoformat()}",
                     f"Weeks: {cfg.weeks}"])
        wr.writerow([])
        # date row + weekday row
        wr.writerow(["Officer"] + [cfg.date_for(t).strftime("%d %b") for t in range(T)])
        wr.writerow([""] + [WEEKDAY_NAMES[cfg.date_for(t).weekday()] for t in range(T)])
        for o, name in enumerate(cfg.officers):
            wr.writerow([name] + [_cell_text(res, o, t) for t in range(T)])
        wr.writerow([])
        wr.writerow(["Summary", "Total Hours", "Day shifts", "Night shifts",
                     "Day-work shifts", "Weekend duties"])
        for o, name in enumerate(cfg.officers):
            wr.writerow([name, res.hours_for(o), res.count_for(o, DAY),
                         res.count_for(o, NIGHT), res.count_for(o, DAYWORK),
                         res.weekend_count_for(o)])
        wr.writerow([])
        wr.writerow(["SCHEDULING RULES"])
        for line in rules_text(cfg):
            wr.writerow([line])


def export_xlsx(res: ScheduleResult, path: str):
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")
    cfg = res.config
    T = cfg.num_days
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DO Schedule"

    fills = {
        DAY: PatternFill("solid", fgColor="C6E0B4"),      # green
        NIGHT: PatternFill("solid", fgColor="BDD7EE"),    # blue
        DAYWORK: PatternFill("solid", fgColor="FFE699"),  # amber
        OFF: PatternFill("solid", fgColor="F2F2F2"),      # light grey
    }
    weekend_hdr = PatternFill("solid", fgColor="FCE4D6")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    base_font = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "DUTY OFFICER SCHEDULE"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws["A2"] = f"Start {cfg.start_date.strftime('%d %b %Y')}  •  {cfg.weeks} weeks"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)

    r0 = 4  # date header row
    ws.cell(r0, 1, "Officer").font = hdr_font
    ws.cell(r0, 1).fill = hdr_fill
    ws.cell(r0 + 1, 1, "").fill = hdr_fill
    for t in range(T):
        c = 2 + t
        dc = ws.cell(r0, c, cfg.date_for(t).strftime("%d %b"))
        wd = ws.cell(r0 + 1, c, WEEKDAY_NAMES[cfg.date_for(t).weekday()])
        for cell in (dc, wd):
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border
            cell.fill = weekend_hdr if cfg.date_for(t).weekday() >= 5 else hdr_fill

    for o, name in enumerate(cfg.officers):
        r = r0 + 2 + o
        nm = ws.cell(r, 1, name)
        nm.font = bold
        nm.border = border
        for t in range(T):
            code = res.grid[o][t]
            cell = ws.cell(r, 2 + t, _cell_text(res, o, t))
            cell.fill = fills[code]
            cell.font = base_font
            cell.alignment = center
            cell.border = border

    # Summary block
    sr = r0 + 2 + len(cfg.officers) + 2
    headers = ["Officer", "Total Hours", "Day", "Night", "Day-work", "Weekend duties"]
    for i, h in enumerate(headers):
        cell = ws.cell(sr, 1 + i, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    for o, name in enumerate(cfg.officers):
        r = sr + 1 + o
        vals = [name, res.hours_for(o), res.count_for(o, DAY), res.count_for(o, NIGHT),
                res.count_for(o, DAYWORK), res.weekend_count_for(o)]
        for i, v in enumerate(vals):
            cell = ws.cell(r, 1 + i, v)
            cell.font = bold if i == 0 else base_font
            cell.alignment = center
            cell.border = border

    # Legend
    lr = sr + len(cfg.officers) + 2
    ws.cell(lr, 1, "Legend:").font = bold
    legend = [(cfg.day_shift.label, DAY), (cfg.night_shift.label, NIGHT),
              (cfg.daywork_shift.label, DAYWORK), ("Off", OFF)]
    for i, (lab, code) in enumerate(legend):
        cell = ws.cell(lr, 2 + i, lab)
        cell.fill = fills[code]
        cell.alignment = center
        cell.border = border
        cell.font = base_font

    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 16
    for t in range(T):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + t)].width = 11

    # ---- Rules sheet ----
    rs = wb.create_sheet("Scheduling Rules")
    rs["A1"] = "SCHEDULING RULES"
    rs["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    rs["A2"] = f"Applies to roster starting {cfg.start_date.strftime('%d %b %Y')}"
    rs["A2"].font = Font(name="Arial", size=10, italic=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for i, line in enumerate(rules_text(cfg)):
        cell = rs.cell(4 + i, 1, line)
        cell.font = base_font
        cell.alignment = wrap
    rs.column_dimensions["A"].width = 110
    wb.save(path)


# ----------------------------------------------------------------------------
# GUI (Tkinter)
# ----------------------------------------------------------------------------
def launch_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    CODE_COLORS = {
        DAY: ("#1b5e20", "#c6e0b4"),
        NIGHT: ("#0d3b66", "#bdd7ee"),
        DAYWORK: ("#7a4f00", "#ffe699"),
        OFF: ("#888888", "#f2f2f2"),
    }

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Duty Officer Schedule Planner")
            self.geometry("1280x780")
            self.minsize(1000, 640)
            self.result = None
            self.cfg = ScheduleConfig()
            self.name_vars = []

            self._build_styles()
            self._build_layout()
            self._rebuild_name_fields()
            if not HAVE_ORTOOLS:
                self.after(300, lambda: messagebox.showwarning(
                    "Missing dependency",
                    "Google OR-Tools is required to generate rosters.\n\n"
                    "Install it from a terminal:\n    pip install ortools"))

        # ---- styling ----
        def _build_styles(self):
            self.configure(bg="#f4f5f7")
            s = ttk.Style(self)
            try:
                s.theme_use("clam")
            except Exception:
                pass
            s.configure("TFrame", background="#f4f5f7")
            s.configure("Card.TFrame", background="#ffffff", relief="flat")
            s.configure("TLabel", background="#f4f5f7", font=("Segoe UI", 10))
            s.configure("Card.TLabel", background="#ffffff", font=("Segoe UI", 10))
            s.configure("Head.TLabel", background="#f4f5f7",
                        font=("Segoe UI", 11, "bold"), foreground="#1f3864")
            s.configure("Title.TLabel", background="#f4f5f7",
                        font=("Segoe UI", 17, "bold"), foreground="#1f3864")
            s.configure("Accent.TButton", font=("Segoe UI", 10, "bold"))
            s.configure("TButton", font=("Segoe UI", 10))
            s.configure("TCheckbutton", background="#ffffff", font=("Segoe UI", 10))

        # ---- layout ----
        def _build_layout(self):
            top = ttk.Frame(self)
            top.pack(fill="x", padx=16, pady=(12, 4))
            ttk.Label(top, text="Duty Officer Schedule Planner", style="Title.TLabel").pack(side="left")
            self.status = ttk.Label(top, text="Ready.", style="Head.TLabel")
            self.status.pack(side="right")

            body = ttk.Frame(self)
            body.pack(fill="both", expand=True, padx=16, pady=8)

            # ---- left settings (scrollable) ----
            left_outer = ttk.Frame(body)
            left_outer.pack(side="left", fill="y")
            lcan = tk.Canvas(left_outer, width=330, bg="#f4f5f7",
                             highlightthickness=0)
            lscroll = ttk.Scrollbar(left_outer, orient="vertical", command=lcan.yview)
            self.settings = ttk.Frame(lcan)
            self.settings.bind("<Configure>",
                               lambda e: lcan.configure(scrollregion=lcan.bbox("all")))
            lcan.create_window((0, 0), window=self.settings, anchor="nw")
            lcan.configure(yscrollcommand=lscroll.set)
            lcan.pack(side="left", fill="y")
            lscroll.pack(side="right", fill="y")
            self._build_settings(self.settings)

            # ---- right (schedule + summary) ----
            right = ttk.Frame(body)
            right.pack(side="left", fill="both", expand=True, padx=(14, 0))

            ttk.Label(right, text="Roster", style="Head.TLabel").pack(anchor="w")
            grid_wrap = ttk.Frame(right, style="Card.TFrame")
            grid_wrap.pack(fill="both", expand=True, pady=(4, 8))
            self.grid_canvas = tk.Canvas(grid_wrap, bg="#ffffff", highlightthickness=0)
            gx = ttk.Scrollbar(grid_wrap, orient="horizontal", command=self.grid_canvas.xview)
            gy = ttk.Scrollbar(grid_wrap, orient="vertical", command=self.grid_canvas.yview)
            self.grid_inner = ttk.Frame(self.grid_canvas, style="Card.TFrame")
            self.grid_inner.bind("<Configure>",
                                 lambda e: self.grid_canvas.configure(
                                     scrollregion=self.grid_canvas.bbox("all")))
            self.grid_canvas.create_window((0, 0), window=self.grid_inner, anchor="nw")
            self.grid_canvas.configure(xscrollcommand=gx.set, yscrollcommand=gy.set)
            self.grid_canvas.grid(row=0, column=0, sticky="nsew")
            gy.grid(row=0, column=1, sticky="ns")
            gx.grid(row=1, column=0, sticky="ew")
            grid_wrap.rowconfigure(0, weight=1)
            grid_wrap.columnconfigure(0, weight=1)

            nb = ttk.Notebook(right)
            nb.pack(fill="x", pady=(4, 0))

            sum_tab = ttk.Frame(nb)
            nb.add(sum_tab, text="Per-officer summary")
            self.summary = ttk.Treeview(
                sum_tab, columns=("hours", "day", "night", "dw", "wknd"),
                show="tree headings", height=7)
            self.summary.heading("#0", text="Officer")
            for col, txt, wdt in [("hours", "Total hrs", 80), ("day", "Day", 60),
                                   ("night", "Night", 60), ("dw", "Day-work", 80),
                                   ("wknd", "Weekend", 80)]:
                self.summary.heading(col, text=txt)
                self.summary.column(col, width=wdt, anchor="center")
            self.summary.column("#0", width=140)
            self.summary.pack(fill="x")

            rules_tab = ttk.Frame(nb)
            nb.add(rules_tab, text="Scheduling rules")
            self.rules_box = tk.Text(rules_tab, height=8, wrap="word",
                                     font=("Segoe UI", 9), bg="#ffffff",
                                     relief="flat", padx=8, pady=6)
            rb_scroll = ttk.Scrollbar(rules_tab, orient="vertical",
                                      command=self.rules_box.yview)
            self.rules_box.configure(yscrollcommand=rb_scroll.set, state="disabled")
            rb_scroll.pack(side="right", fill="y")
            self.rules_box.pack(fill="both", expand=True)
            self._refresh_rules()

        def _section(self, parent, title):
            ttk.Label(parent, text=title, style="Head.TLabel").pack(
                anchor="w", pady=(12, 2))
            card = ttk.Frame(parent, style="Card.TFrame", padding=10)
            card.pack(fill="x")
            return card

        def _build_settings(self, p):
            # Officers
            c = self._section(p, "Duty Officers")
            row = ttk.Frame(c, style="Card.TFrame"); row.pack(fill="x")
            ttk.Label(row, text="Number of officers:", style="Card.TLabel").pack(side="left")
            self.n_off = tk.IntVar(value=len(self.cfg.officers))
            sp = ttk.Spinbox(row, from_=2, to=8, width=5, textvariable=self.n_off,
                             command=self._rebuild_name_fields)
            sp.pack(side="left", padx=6)
            self.names_frame = ttk.Frame(c, style="Card.TFrame")
            self.names_frame.pack(fill="x", pady=(6, 0))

            # Period
            c = self._section(p, "Period")
            r1 = ttk.Frame(c, style="Card.TFrame"); r1.pack(fill="x")
            ttk.Label(r1, text="Weeks:", style="Card.TLabel").pack(side="left")
            self.weeks = tk.IntVar(value=self.cfg.weeks)
            ttk.Spinbox(r1, from_=1, to=26, width=5, textvariable=self.weeks).pack(side="left", padx=6)
            r2 = ttk.Frame(c, style="Card.TFrame"); r2.pack(fill="x", pady=(6, 0))
            ttk.Label(r2, text="Start (Mon, YYYY-MM-DD):", style="Card.TLabel").pack(side="left")
            self.start = tk.StringVar(value=self.cfg.start_date.isoformat())
            ttk.Entry(r2, width=12, textvariable=self.start).pack(side="left", padx=6)

            # Day-work
            c = self._section(p, "8h Day-work (Mon-Fri, rotated weekly)")
            r = ttk.Frame(c, style="Card.TFrame"); r.pack(fill="x")
            ttk.Label(r, text="Officers per week:", style="Card.TLabel").pack(side="left")
            self.dw = tk.IntVar(value=self.cfg.daywork_per_week)
            ttk.Spinbox(r, from_=0, to=2, width=5, textvariable=self.dw).pack(side="left", padx=6)

            # Shift definitions
            c = self._section(p, "Shift Definitions")
            self.shift_vars = {}
            for key, sd in [("day", self.cfg.day_shift), ("night", self.cfg.night_shift),
                            ("dw", self.cfg.daywork_shift)]:
                fr = ttk.Frame(c, style="Card.TFrame"); fr.pack(fill="x", pady=2)
                lab = tk.StringVar(value=sd.label)
                tt = tk.StringVar(value=sd.time_text)
                hh = tk.IntVar(value=sd.hours)
                ttk.Entry(fr, width=9, textvariable=lab).pack(side="left")
                ttk.Entry(fr, width=11, textvariable=tt).pack(side="left", padx=4)
                ttk.Spinbox(fr, from_=1, to=24, width=4, textvariable=hh).pack(side="left")
                ttk.Label(fr, text="h", style="Card.TLabel").pack(side="left")
                self.shift_vars[key] = (lab, tt, hh)

            # Block rules
            c = self._section(p, "Block Rules (12h watches)")
            r = ttk.Frame(c, style="Card.TFrame"); r.pack(fill="x")
            ttk.Label(r, text="Min consecutive:", style="Card.TLabel").pack(side="left")
            self.minc = tk.IntVar(value=self.cfg.min_consecutive)
            ttk.Spinbox(r, from_=1, to=5, width=4, textvariable=self.minc).pack(side="left", padx=4)
            ttk.Label(r, text="Max:", style="Card.TLabel").pack(side="left")
            self.maxc = tk.IntVar(value=self.cfg.max_consecutive)
            ttk.Spinbox(r, from_=2, to=7, width=4, textvariable=self.maxc).pack(side="left", padx=4)
            r2 = ttk.Frame(c, style="Card.TFrame"); r2.pack(fill="x", pady=(6, 0))
            ttk.Label(r2, text="Rest days after block:", style="Card.TLabel").pack(side="left")
            self.rest = tk.IntVar(value=self.cfg.rest_days_after_block)
            ttk.Spinbox(r2, from_=1, to=4, width=4, textvariable=self.rest).pack(side="left", padx=4)
            r3 = ttk.Frame(c, style="Card.TFrame"); r3.pack(fill="x", pady=(6, 0))
            ttk.Label(r3, text="Off days when day<->night:", style="Card.TLabel").pack(side="left")
            self.swdn = tk.IntVar(value=self.cfg.switch_daynight_off_days)
            ttk.Spinbox(r3, from_=0, to=5, width=4, textvariable=self.swdn).pack(side="left", padx=4)
            r4 = ttk.Frame(c, style="Card.TFrame"); r4.pack(fill="x", pady=(6, 0))
            ttk.Label(r4, text="Off days when 12h<->8h:", style="Card.TLabel").pack(side="left")
            self.swwh = tk.IntVar(value=self.cfg.switch_workhours_off_days)
            ttk.Spinbox(r4, from_=0, to=5, width=4, textvariable=self.swwh).pack(side="left", padx=4)
            self.enf = tk.BooleanVar(value=self.cfg.enforce_min_consecutive)
            ttk.Checkbutton(c, text="Strictly enforce minimum consecutive",
                            variable=self.enf).pack(anchor="w", pady=(6, 0))

            # Advanced fairness weights
            c = self._section(p, "Fairness Weights (advanced)")
            self.weights = {}
            for key, label, default in [("hours", "Equal hours", self.cfg.w_hours),
                                        ("dn_self", "Own day/night balance", self.cfg.w_daynight_self),
                                        ("dn_spread", "Day/night across team", self.cfg.w_daynight_spread),
                                        ("wknd", "Equal weekends", self.cfg.w_weekend),
                                        ("dwrot", "Day-work rotation", self.cfg.w_daywork_rotation)]:
                fr = ttk.Frame(c, style="Card.TFrame"); fr.pack(fill="x", pady=1)
                ttk.Label(fr, text=label, style="Card.TLabel", width=22).pack(side="left")
                v = tk.IntVar(value=default)
                ttk.Spinbox(fr, from_=0, to=500, width=6, textvariable=v).pack(side="left")
                self.weights[key] = v
            fr = ttk.Frame(c, style="Card.TFrame"); fr.pack(fill="x", pady=(6, 0))
            ttk.Label(fr, text="Solver time limit (s):", style="Card.TLabel").pack(side="left")
            self.tlimit = tk.IntVar(value=self.cfg.time_limit_s)
            ttk.Spinbox(fr, from_=2, to=120, width=6, textvariable=self.tlimit).pack(side="left", padx=4)

            # Actions
            c = self._section(p, "Actions")
            ttk.Button(c, text="Generate Roster", style="Accent.TButton",
                       command=lambda: self.generate(new_seed=True)).pack(fill="x", pady=2)
            ttk.Button(c, text="Re-generate (new variant)",
                       command=lambda: self.generate(new_seed=True)).pack(fill="x", pady=2)
            er = ttk.Frame(c, style="Card.TFrame"); er.pack(fill="x", pady=(6, 0))
            ttk.Button(er, text="Export Excel", command=self.export_excel).pack(
                side="left", expand=True, fill="x", padx=(0, 3))
            ttk.Button(er, text="Export CSV", command=self.export_csv_action).pack(
                side="left", expand=True, fill="x", padx=(3, 0))
            sr = ttk.Frame(c, style="Card.TFrame"); sr.pack(fill="x", pady=(4, 0))
            ttk.Button(sr, text="Save", command=self.save_action).pack(
                side="left", expand=True, fill="x", padx=(0, 3))
            ttk.Button(sr, text="Load", command=self.load_action).pack(
                side="left", expand=True, fill="x", padx=(3, 0))

        def _rebuild_name_fields(self):
            for ch in self.names_frame.winfo_children():
                ch.destroy()
            target = self.n_off.get()
            old = [v.get() for v in self.name_vars]
            self.name_vars = []
            for i in range(target):
                default = old[i] if i < len(old) else f"DO {i+1}"
                v = tk.StringVar(value=default)
                fr = ttk.Frame(self.names_frame, style="Card.TFrame"); fr.pack(fill="x", pady=1)
                ttk.Label(fr, text=f"{i+1}.", style="Card.TLabel", width=3).pack(side="left")
                ttk.Entry(fr, textvariable=v).pack(side="left", fill="x", expand=True)
                self.name_vars.append(v)

        # ---- gather config ----
        def _gather_config(self):
            try:
                sd = dt.date.fromisoformat(self.start.get().strip())
            except Exception:
                raise ValueError("Start date must be YYYY-MM-DD (e.g. 2026-06-15).")
            if sd.weekday() != 0:
                sd = sd - dt.timedelta(days=sd.weekday())  # snap back to Monday
                self.start.set(sd.isoformat())
            names = [v.get().strip() or f"DO {i+1}" for i, v in enumerate(self.name_vars)]
            sv = self.shift_vars
            cfg = ScheduleConfig(
                officers=names,
                weeks=max(1, self.weeks.get()),
                start_date=sd,
                day_shift=ShiftDef(sv["day"][0].get(), sv["day"][1].get(), sv["day"][2].get()),
                night_shift=ShiftDef(sv["night"][0].get(), sv["night"][1].get(), sv["night"][2].get()),
                daywork_shift=ShiftDef(sv["dw"][0].get(), sv["dw"][1].get(), sv["dw"][2].get()),
                daywork_per_week=self.dw.get(),
                min_consecutive=self.minc.get(),
                max_consecutive=self.maxc.get(),
                rest_days_after_block=self.rest.get(),
                switch_daynight_off_days=self.swdn.get(),
                switch_workhours_off_days=self.swwh.get(),
                enforce_min_consecutive=self.enf.get(),
                w_hours=self.weights["hours"].get(),
                w_daynight_self=self.weights["dn_self"].get(),
                w_daynight_spread=self.weights["dn_spread"].get(),
                w_weekend=self.weights["wknd"].get(),
                w_daywork_rotation=self.weights["dwrot"].get(),
                time_limit_s=self.tlimit.get(),
            )
            return cfg

        def _refresh_rules(self):
            try:
                cfg = self._gather_config()
            except Exception:
                cfg = self.cfg
            self.rules_box.configure(state="normal")
            self.rules_box.delete("1.0", "end")
            self.rules_box.insert("end", "\n".join(rules_text(cfg)))
            self.rules_box.configure(state="disabled")

        def _apply_config_to_ui(self, cfg):
            self.cfg = cfg
            self.n_off.set(len(cfg.officers))
            self._rebuild_name_fields()
            for i, v in enumerate(self.name_vars):
                if i < len(cfg.officers):
                    v.set(cfg.officers[i])
            self.weeks.set(cfg.weeks)
            self.start.set(cfg.start_date.isoformat())
            self.dw.set(cfg.daywork_per_week)
            for key, sd in [("day", cfg.day_shift), ("night", cfg.night_shift),
                            ("dw", cfg.daywork_shift)]:
                lab, tt, hh = self.shift_vars[key]
                lab.set(sd.label); tt.set(sd.time_text); hh.set(sd.hours)
            self.minc.set(cfg.min_consecutive)
            self.maxc.set(cfg.max_consecutive)
            self.rest.set(cfg.rest_days_after_block)
            self.swdn.set(cfg.switch_daynight_off_days)
            self.swwh.set(cfg.switch_workhours_off_days)
            self.enf.set(cfg.enforce_min_consecutive)
            self.weights["hours"].set(cfg.w_hours)
            self.weights["dn_self"].set(cfg.w_daynight_self)
            self.weights["dn_spread"].set(cfg.w_daynight_spread)
            self.weights["wknd"].set(cfg.w_weekend)
            self.weights["dwrot"].set(cfg.w_daywork_rotation)
            self.tlimit.set(cfg.time_limit_s)
            self._refresh_rules()

        # ---- generate ----
        def generate(self, new_seed=True):
            try:
                cfg = self._gather_config()
            except ValueError as e:
                messagebox.showerror("Invalid settings", str(e))
                return
            self.cfg = cfg
            self.status.config(text="Solving…")
            self.update_idletasks()
            res = generate_schedule(cfg, seed=random.randint(1, 2_000_000_000) if new_seed else 1)
            self.result = res
            if not res.ok:
                self.status.config(text=res.status)
                messagebox.showwarning("No roster", res.message)
                self._clear_grid()
                return
            probs = validate(res)
            extra = "" if not probs else f"  ({len(probs)} soft issues)"
            self.status.config(text=f"{res.status}{extra}")
            self._render(res)

        def _clear_grid(self):
            for ch in self.grid_inner.winfo_children():
                ch.destroy()
            for it in self.summary.get_children():
                self.summary.delete(it)

        def _render(self, res):
            self._clear_grid()
            cfg = res.config
            T = cfg.num_days
            hdr = {"bg": "#1f3864", "fg": "white", "font": ("Segoe UI", 9, "bold")}
            wkhdr = {"bg": "#8a5a44", "fg": "white", "font": ("Segoe UI", 9, "bold")}

            tk.Label(self.grid_inner, text="Officer", **hdr, padx=8, pady=4,
                     borderwidth=1, relief="solid").grid(row=0, column=0, rowspan=2, sticky="nsew")
            for t in range(T):
                wknd = cfg.date_for(t).weekday() >= 5
                style = wkhdr if wknd else hdr
                tk.Label(self.grid_inner, text=cfg.date_for(t).strftime("%d %b"),
                         **style, padx=4, pady=2, borderwidth=1, relief="solid"
                         ).grid(row=0, column=1 + t, sticky="nsew")
                tk.Label(self.grid_inner, text=WEEKDAY_NAMES[cfg.date_for(t).weekday()],
                         **style, padx=4, pady=2, borderwidth=1, relief="solid"
                         ).grid(row=1, column=1 + t, sticky="nsew")

            for o, name in enumerate(cfg.officers):
                tk.Label(self.grid_inner, text=name, font=("Segoe UI", 9, "bold"),
                         bg="#eef1f6", padx=8, pady=3, borderwidth=1, relief="solid"
                         ).grid(row=2 + o, column=0, sticky="nsew")
                for t in range(T):
                    code = res.grid[o][t]
                    fg, bg = CODE_COLORS[code]
                    tk.Label(self.grid_inner, text=_cell_text(res, o, t),
                             fg=fg, bg=bg, font=("Segoe UI", 8), padx=2, pady=3,
                             borderwidth=1, relief="solid"
                             ).grid(row=2 + o, column=1 + t, sticky="nsew")

            for o, name in enumerate(cfg.officers):
                self.summary.insert("", "end", text=name,
                                    values=(res.hours_for(o), res.count_for(o, DAY),
                                            res.count_for(o, NIGHT), res.count_for(o, DAYWORK),
                                            res.weekend_count_for(o)))
            self._refresh_rules()

        # ---- export ----
        def export_excel(self):
            if not self.result or not self.result.ok:
                messagebox.showinfo("Nothing to export", "Generate a roster first.")
                return
            if not HAVE_OPENPYXL:
                messagebox.showerror("Missing dependency",
                                     "openpyxl is required for Excel export.\n\n"
                                     "Install it: pip install openpyxl\n\n"
                                     "You can still export to CSV.")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                initialfile="DO_Schedule.xlsx")
            if not path:
                return
            try:
                export_xlsx(self.result, path)
                messagebox.showinfo("Exported", f"Saved:\n{path}")
            except Exception as e:
                messagebox.showerror("Export failed", str(e))

        def export_csv_action(self):
            if not self.result or not self.result.ok:
                messagebox.showinfo("Nothing to export", "Generate a roster first.")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".csv", filetypes=[("CSV", "*.csv")],
                initialfile="DO_Schedule.csv")
            if not path:
                return
            try:
                export_csv(self.result, path)
                messagebox.showinfo("Exported", f"Saved:\n{path}")
            except Exception as e:
                messagebox.showerror("Export failed", str(e))

        def save_action(self):
            try:
                cfg = self._gather_config()
            except ValueError as e:
                messagebox.showerror("Invalid settings", str(e))
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".json", filetypes=[("DO schedule file", "*.json")],
                initialfile="DO_Schedule_Session.json")
            if not path:
                return
            try:
                save_session(path, cfg, self.result)
                messagebox.showinfo("Saved", f"Settings and roster saved:\n{path}")
            except Exception as e:
                messagebox.showerror("Save failed", str(e))

        def load_action(self):
            path = filedialog.askopenfilename(
                filetypes=[("DO schedule file", "*.json"), ("All files", "*.*")])
            if not path:
                return
            try:
                cfg, res = load_session(path)
            except Exception as e:
                messagebox.showerror("Load failed", str(e))
                return
            self._apply_config_to_ui(cfg)
            self.result = res
            if res and res.ok:
                self._render(res)
                self.status.config(text="Loaded saved roster.")
            else:
                self._clear_grid()
                self.status.config(text="Loaded settings (no saved roster).")

    App().mainloop()


# ----------------------------------------------------------------------------
# Self-test (headless)
# ----------------------------------------------------------------------------
def _selftest():
    scenarios = [
        ("4 officers, no day-work, 4 wks", dict(officers=["A", "B", "C", "D"], weeks=4, daywork_per_week=0)),
        ("5 officers, 1 day-work, 4 wks", dict(officers=["A", "B", "C", "D", "E"], weeks=4, daywork_per_week=1)),
        ("6 officers, 2 day-work, 4 wks", dict(officers=["A", "B", "C", "D", "E", "F"], weeks=4, daywork_per_week=2)),
        ("4 officers, no day-work, 6 wks", dict(officers=["A", "B", "C", "D"], weeks=6, daywork_per_week=0)),
        ("6 officers, no day-work, 8 wks", dict(officers=["A", "B", "C", "D", "E", "F"], weeks=8, daywork_per_week=0)),
    ]
    allok = True
    for name, kw in scenarios:
        cfg = ScheduleConfig(**kw)
        res = generate_schedule(cfg, seed=42)
        if not res.ok:
            print(f"[FAIL] {name}: {res.status} - {res.message.splitlines()[0]}")
            allok = False
            continue
        probs = validate(res)
        hrs = [res.hours_for(o) for o in range(len(cfg.officers))]
        days = [res.count_for(o, DAY) for o in range(len(cfg.officers))]
        nights = [res.count_for(o, NIGHT) for o in range(len(cfg.officers))]
        wknd = [res.weekend_count_for(o) for o in range(len(cfg.officers))]
        tag = "OK " if not probs else "BAD"
        if probs:
            allok = False
        print(f"[{tag}] {name}: {res.status}")
        print(f"        hours={hrs} (spread {max(hrs)-min(hrs)})  days={days}  nights={nights}  weekend={wknd}")
        if probs:
            for p in probs[:8]:
                print(f"        ! {p}")
    print("\nALL SCENARIOS VALID" if allok else "\nSOME SCENARIOS FAILED")
    return allok


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        ok = _selftest()
        sys.exit(0 if ok else 1)
    else:
        launch_gui()
